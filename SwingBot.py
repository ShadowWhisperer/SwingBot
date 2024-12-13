#!/usr/bin/python3

#
# Creator: ShadowWhisperer
#  Github: https://github.com/ShadowWhisperer
# Created: 12/10/2024
# Updated: 12/12/2024
#

import csv
import os
import sys
import yfinance as yf
from openpyxl import load_workbook
import pandas as pd
from ta.momentum import RSIIndicator
from ta.trend import EMAIndicator, MACD
from datetime import datetime, timedelta

# Parameters
#####################################
rsi_min = 27               # Minimum RSI threshold       27
rsi_max = 70               # Maximum RSI threshold       70
price_min = 2              # Minimum $                    2
price_max = 20             # Max $                       20
ema_days = 14              # Days for EMA to check       14
macd_fast = 12             # MACD Fast period            12
macd_slow = 26             # MACD Slow period            26
macd_signal = 9            # MACD Signal period           9
relative_vol_min = 0.6     # Relative volume min          0.6
avg_vol_min = 500_000      # Average volume min           500_000
yearly_low_threshold = 10  # Percentage above the yearly low to consider "low"
recent_days = 4            # Number of recent days to analyze
DEBUG = False              # Show debug info
#####################################


# Color Variables
blue = "\033[94m"
green = "\033[92m"
red = "\033[91m"
white = "\033[0m"


def clear_screen():
    if os.name == 'nt':
        _ = os.system('cls')
    else:
        _ = os.system('clear')

if not os.path.exists('tickers.txt'):
    print("Error: 'tickers.txt' file not found.")
    sys.exit()
clear_screen()

# Print Stats
print(f"\n  Analysing {blue}{sum(1 for _ in open('tickers.txt'))}{white} Stocks @ {green}${price_min}{white} -> {green}${price_max}{white}\n")


# Date range utilities
def get_date_range():
    end_date = datetime.today().strftime('%Y-%m-%d')
    start_date = (datetime.today() - timedelta(days=90)).strftime('%Y-%m-%d')
    return start_date, end_date

def get_date_ranges():
    end_date = datetime.today().strftime('%Y-%m-%d')
    start_date_year = (datetime.today() - timedelta(days=365)).strftime('%Y-%m-%d')
    start_date_recent = (datetime.today() - timedelta(days=recent_days)).strftime('%Y-%m-%d')
    return start_date_year, start_date_recent, end_date

# Technical analysis
def fetch_technical_data(symbol):
    start_date, end_date = get_date_range()
    data = yf.download(symbol, start=start_date, end=end_date, interval="1d", progress=False)

    if data.empty:
        print(f"{red}No data for {symbol}{white}")
        return None

    # Handle multi-index columns
    if isinstance(data.columns, pd.MultiIndex):
        data.columns = data.columns.droplevel(0)
        data.columns = ["Open", "High", "Low", "Close", "Adj Close", "Volume"]

    if "Adj Close" not in data.columns or "Volume" not in data.columns:
        return None

    try:
        close_series = data["Adj Close"]
        if close_series.ndim != 1:
            print(f"{red}Error - {symbol}: Data must be 1-dimensional{white}")
            return None

        rsi_indicator = RSIIndicator(close_series)
        ema_indicator = EMAIndicator(close_series, window=ema_days)
        macd_indicator = MACD(close_series, window_slow=macd_slow, window_fast=macd_fast, window_sign=macd_signal)

        data["EMA"] = ema_indicator.ema_indicator()
        data["MACD"] = macd_indicator.macd()
        data["MACD_Signal"] = macd_indicator.macd_signal()
        data["MACD_Diff"] = macd_indicator.macd_diff()
        data["RSI"] = rsi_indicator.rsi()
        data["AvgVolume"] = data["Volume"].rolling(window=10).mean()
        data["RelVolume"] = data["Volume"] / data["AvgVolume"]

        latest_data = data.iloc[-1]
        if (
            rsi_min < latest_data["RSI"] < rsi_max
            and price_min < latest_data["Close"] < price_max
            and latest_data["AvgVolume"] > avg_vol_min
            and latest_data["RelVolume"] > relative_vol_min
            and latest_data["MACD_Diff"] > 0
        ):
            return {
                "symbol": symbol,
                "rsi": latest_data["RSI"],
                "price": latest_data["Close"],
                "rel_volume": latest_data["RelVolume"],
                "avg_volume": latest_data["AvgVolume"],
                "ema": latest_data["EMA"],
                "macd": latest_data["MACD"],
                "macd_signal": latest_data["MACD_Signal"],
                "macd_diff": latest_data["MACD_Diff"],
            }
    except Exception as e:
        print(f"{red}Error - {symbol}: {e}{white}")
        return None

# Fundamental analysis
def get_company_overview(ticker):
    stock = yf.Ticker(ticker)
    try:
        info = stock.info
        return {
            'Name': info.get('longName', 'N/A'),
            'Sector': info.get('sector', 'N/A'),
            'MarketCap': info.get('marketCap', 0),
            'PERatio': info.get('trailingPE', 0),
            'NetIncome': info.get('netIncomeToCommon', 0),
            'RevenueGrowth': info.get('revenueGrowth', None),
            'DebtToEquity': info.get('debtToEquity', None)
        }
    except Exception as e:
        print(f"{red}Error fetching data - {ticker}: {e}{white}")
        return {}

# Yearly low analysis
def analyze_yearly_low(symbol):
    start_year, start_recent, end_date = get_date_ranges()
    try:
        data = yf.download(symbol, start=start_year, end=end_date, progress=False)

        if data.empty:
            print(f"No data for {symbol}")
            return None

        yearly_low = data['Low'].min()
        yearly_low = yearly_low.item() if hasattr(yearly_low, 'item') else yearly_low

        recent_data = data[data.index >= start_recent]

        if recent_data.empty:
            print(f"No recent data for {symbol}")
            return None

        current_price = recent_data['Close'].iloc[-1]
        current_price = current_price.item() if hasattr(current_price, 'item') else current_price

        percentage_above_low = ((current_price - yearly_low) / yearly_low) * 100

        # Debugging logs
        if DEBUG:
            print(f" [{symbol}] - Yearly Low: {yearly_low}, Current Price: {current_price}, "
                  f"Percentage Above Low: {percentage_above_low:.2f}%")

        if percentage_above_low <= yearly_low_threshold:
            return {
                "symbol": symbol,
                "current_price": current_price,
                "yearly_low": yearly_low,
                "percentage_above_low": percentage_above_low
            }
        else:
            return None

    except Exception as e:
        print(f"Error analyzing {symbol}: {e}")
        return None

# Evaluate stability
def evaluate_company_stability(company_data):
    market_cap = company_data.get('MarketCap', 0)
    pe_ratio = company_data.get('PERatio', 0)
    net_income = company_data.get('NetIncome', 0)
    revenue_growth = company_data.get('RevenueGrowth', None)
    debt_to_equity = company_data.get('DebtToEquity', None)

    stable = True

    if market_cap < 10_000_000_000:
        stable = False
    if pe_ratio and pe_ratio > 30:
        stable = False
    if net_income and net_income < 0:
        stable = False
    if revenue_growth is not None and revenue_growth < 0.05:
        stable = False
    if debt_to_equity is not None and debt_to_equity > 1.0:
        stable = False

    return stable

# Combine analyses
def combine_analysis(tickers):
    results = []
    for ticker in tickers:
        technicals = fetch_technical_data(ticker)
        yearly_low_info = analyze_yearly_low(ticker)
        if technicals:
            fundamentals = get_company_overview(ticker)
            stability = evaluate_company_stability(fundamentals)
            technicals.update(fundamentals)
            technicals['stable'] = stability
            if yearly_low_info:
                technicals.update(yearly_low_info)
            results.append(technicals)
    return results

# Main execution
tickers = [line.strip() for line in open('tickers.txt')]
candidates = combine_analysis(tickers)
print("\n-------------------------------------")
print("         [Analysis Results]")
print("-------------------------------------\n")
data_list = []
if candidates:
    for candidate in candidates:
        # Show results
        print(f"{blue}{candidate['symbol']}{white}")
        if 'yearly_low' in candidate:
            print(f"  Yearly Low: ${candidate['yearly_low']:.2f}")
            print(f"  Current Price: ${candidate['current_price']:.2f}")
            print(f"  Percentage Above Low: {candidate['percentage_above_low']:.2f}%")
        
        # Prepare data for each candidate
        data = {
            'symbol': candidate['symbol'],
            'Name': candidate.get('Name', 'N/A'),
            'Sector': candidate.get('Sector', 'N/A'),
            'price': f"${candidate['price']:.2f}",
            'rsi': f"{candidate['rsi']:.2f}",
            'rel_volume': f"{candidate['rel_volume']:.2f}",
            'avg_volume': f"{candidate['avg_volume']:.2f}",
            'ema': f"${candidate['ema']:.2f}",
            'macd': f"{candidate['macd']:.2f}",
            'macd_signal': f"{candidate['macd_signal']:.2f}",
            'macd_diff': f"{candidate['macd_diff']:.2f}",
            'MarketCap': candidate.get('MarketCap', 'N/A'),
            'PERatio': candidate.get('PERatio', 'N/A'),
            'NetIncome': candidate.get('NetIncome', 'N/A'),
            'RevenueGrowth': candidate.get('RevenueGrowth', 'N/A'),
            'DebtToEquity': candidate.get('DebtToEquity', 'N/A'),
            'yearly_low': f"${candidate['yearly_low']:.2f}" if 'yearly_low' in candidate else 'N/A',
            'current_price': f"${candidate['current_price']:.2f}" if 'current_price' in candidate else 'N/A',
            'percentage_above_low': f"{candidate['percentage_above_low']:.2f}%" if 'percentage_above_low' in candidate else 'N/A',
            'stability': "Stable" if candidate.get('stable', False) else "May be risky"
        }
        data_list.append(data)
else:
    print("No good candidates found.")

#Save data to spreadsheet
df = pd.DataFrame(data_list)
file_name = 'candidate_data.xlsx'
df.to_excel(file_name, index=False)

#Auto-expand columns to fit content
from openpyxl.utils import get_column_letter
workbook = load_workbook(file_name)
sheet = workbook.active
for column in sheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    sheet.column_dimensions[column_letter].width = max_length + 2
workbook.save(file_name)
